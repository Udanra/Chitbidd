from sqlite3 import Timestamp
from tkinter import N
from django.shortcuts import render,redirect
from django.http import HttpResponse,HttpResponseBadRequest

from google.cloud import firestore
from firebase_admin import firestore

import firebase_admin
from firebase_admin import credentials


from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
import os

from openpyxl import Workbook


if not firebase_admin._apps:
    cred = credentials.Certificate('E:/Django/chitbidTask/serviceAccountKey.json')
    firebase_admin.initialize_app(cred)

    db = firestore.client() 



# Start User_Page create user details 
def create_document(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        dob = request.POST.get('dob')
        email = request.POST.get('email')
        gender = request.POST.get('gender')
        phone_no = request.POST.get('phone_no')
        organization = request.POST.get('organization')
        language = request.POST.get('language')
        agent = request.POST.get('agent')
        profile = request.POST.get('profile')
        email_verified = request.POST.get('email_verified')
        uid = request.POST.get('uid')
        url = request.POST.get('url')
        

        db = firestore.Client()
        doc_ref = db.collection('users').document(phone_no)
        doc_ref.set({
            'name': name,
            'dob': dob,
            'email':email,
            'gender':gender,
            'phone_no':phone_no,
            'organization':organization,
            'language':language,
            'agent':agent,
            'profile':profile,
            'email_verified':email_verified,
            'uid':uid,
            'url':url,
        }) 

        return redirect('read_documents')
        
    return render(request, 'create.html')
   
#End User_Page create user details...


# Start User_Page read user details
def read_documents(request):
    db = firestore.Client()
    docs = db.collection('users').get()
 
    
    users = []
    for doc in docs:
        user = doc.to_dict()
        user['id'] = doc.id
        users.append(user)
    return render(request, 'user_read.html', {'users': users})

# End User_Page read user details...


# Start User_Page update user details...
def update_document(request, document_id):
    db = firestore.Client()
    doc_ref = db.collection('users').document(document_id)
    user = doc_ref.get().to_dict()
    
    if request.method == 'POST':
        name = request.POST.get('name')
        dob = request.POST.get('dob')
        email = request.POST.get('email')
        gender = request.POST.get('gender')
        phone_no = request.POST.get('phone_no')
        organization = request.POST.get('organization')
        language = request.POST.get('language')
        agent = request.POST.get('agent')
        profile = request.POST.get('profile')
        email_verified = request.POST.get('email_verified')
        uid = request.POST.get('uid')
        url = request.POST.get('url')
        
        doc_ref.update({
            'name': name,
            'dob': dob,
            'email':email,
            'gender':gender,
            'phone_no':phone_no,
            'organization':organization,
            'language':language,
            'agent':agent,
            'profile':profile,
            'email_verified':email_verified,
            'uid':uid,
            'url':url,
        })
        
        return redirect('read_documents')
    
    return render(request, 'update.html', {'document_id': document_id, 'user':user}) 

# End User_Page update user details...


#Start User_Page delete user details
def delete_document(request, document_id):
    db = firestore.Client()
    doc_ref = db.collection('users').document(document_id)
    doc_ref.delete()
    
    return redirect('read_documents')

# End User_Page delete user details...


# Start Login...
def login(request):
    error_message = None 


    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        
        # Initialize Firestore database
        db = firestore.Client()
        
        # Retrieve the user document from the 'users' collection
        user_doc = db.collection('login').document(username).get()
        
        if user_doc.exists and user_doc.get('password') == password:
            # User authenticated, perform additional actions or redirect to a success page
            return redirect('home')
        else:
            # User authentication failed, display an error message or redirect to a failure page
            error_message = 'Invalid username or password'
          
    
    return render(request, 'login.html',{'error_message':error_message})
# End Login...


# Start Home page...
def home_page(request):
    return render(request, 'dashboard.html')
# End Home page...


# Start User_Page pdf download...
def generate_pdf(document_id):
    print( document_id,  "Generating PDF...")
    db = firestore.Client()
    doc_ref = db.collection('users').document(document_id)
    users = doc_ref.get()
    
    print(users,5456)
   
    if users.exists:
        data = users.to_dict()
        pdf_filename = 'data.pdf'
        pdf_path = os.path.abspath(pdf_filename)
        print("PDF file path:", pdf_path)
        # Generate PDF
      
        c = canvas.Canvas(pdf_filename, pagesize=letter)
        
        # Write data to PDF
        c.drawString(100, 700, f"Name: {data.get('name')}")
        c.drawString(100, 680, f"Date Of Birth: {data.get('dob')}")
        c.drawString(100, 660, f"E-mail: {data.get('email')}")
        c.drawString(100, 640, f"Gender: {data.get('gender')}")
        c.drawString(100, 620, f"Phone_no: {data.get('phone_no')}")
        c.drawString(100, 600, f"Organization: {data.get('organization')}")
        c.drawString(100, 580, f"Language: {data.get('language')}")
        c.drawString(100, 560, f"Agent: {data.get('agent')}")
        c.drawString(100, 540, f"Profile: {data.get('profile')}")
        c.drawString(100, 520, f"E-mail Verified: {data.get('email_verified')}")
        c.drawString(100, 500, f"Uid: {data.get('uid')}")
        c.drawString(100, 480, f"Url: {data.get('url')}")
       
        # Add more fields as necessary
        
        c.save()
        
        return pdf_filename
    
    return None



def download_pdf(request, document_id):
    pdf_filename = generate_pdf(document_id)

    if pdf_filename:
        try:
            with open(pdf_filename, 'rb') as file:
                response = HttpResponse(file.read(), content_type='application/pdf')
                response['Content-Disposition'] = 'attachment; filename=' + os.path.basename(pdf_filename)
                print("Response content size:", len(response.content))  # Add this line
                return response
        except FileNotFoundError:
            return HttpResponse("PDF file not found.")
        except Exception as e:
            return HttpResponse("Error generating PDF: " + str(e))

    return HttpResponse("PDF generation failed.") 

# End User_Page pdf download...

# Start User_Page Excel Sheet Download...

def download_excel(request, document_id):
    # Fetch data from Firestore or any other source
    print(document_id, 456464)
    db = firestore.Client()
    doc_ref = db.collection('users').document(document_id)
    user_doc = doc_ref.get()
    user_data = user_doc.to_dict()

    # Create a new workbook
    workbook = Workbook()

    # Create a new sheet in the workbook
    sheet = workbook.active

    # Add header row
    headers = ['Name', 'Date Of Birth', 'Email', 'Gender', 'Phone_no', 'Organization', 'Language', 'Agent', 'Profile', 'Email_verified','Uid','Url']  # Add more fields as per your Firestore document structure
    sheet.append(headers)

    # Add data row
    row_data = [user_data['name'], user_data['dob'], user_data['email'], user_data['gender'], user_data['phone_no'], user_data['organization'], user_data['language'], user_data['agent'], user_data['profile'], user_data['email_verified'], user_data['uid'], user_data['url']]  # Add more fields as per your Firestore document structure
    sheet.append(row_data)

    # Set the response headers for downloading the file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=table_data.xlsx'

    # Save the workbook to the response
    workbook.save(response)

    return response

# End User_Page Excel Sheet Download...


# Bids Page Related codes here...

# Start Bid_Page Read Function
def bids(request):
    db = firestore.Client()
    docs = db.collection('bids').get()

    bids = []
    for doc in docs:
        bid = doc.to_dict()
        bid['id'] = doc.id
        bids.append(bid)
    return render (request, 'bids.html', {'bids':bids})
#End Bid_Page Read Function

# Start Bid_Page PDF Download
def generate_bid_pdf(document_id):
    print( document_id,  "Generating PDF...")
    db = firestore.Client()
    doc_ref = db.collection('bids').document(document_id)
    bid_users = doc_ref.get()
    
    
   
    if bid_users.exists:
        data = bid_users.to_dict()
        bid_pdf_filename = 'Bid_data.pdf'
        bid_pdf_path = os.path.abspath(bid_pdf_filename)
        print("PDF file path:", bid_pdf_path)
        # Generate PDF
      
        c = canvas.Canvas(bid_pdf_filename, pagesize=letter)
        
        # Write data to PDF
        c.drawString(100, 700, f"Bid_Amount: {data.get('bid_amount')}")
        c.drawString(100, 680, f"Bid Date & Time: {data.get('bid_date')}")
        c.drawString(100, 660, f"Nth_Bid: {data.get('nth_bid')}")
        c.drawString(100, 640, f"Chit_ID: {data.get('chit_id')}")
        c.drawString(100, 620, f"Winner_Phone_no: {data.get('winner_phone_no')}")
        c.drawString(100, 600, f"Number_OF_Bells: {data.get('num_of_bells')}")
        c.drawString(100, 580, f"Bid_at: {data.get('bid_at')}")
        c.drawString(100, 560, f"Ended_at: {data.get('ended_at')}")
        c.drawString(100, 540, f"Is_Active: {data.get('is_active')}")
        c.drawString(100, 520, f"Is_Agent_bid: {data.get('is_agent_bid')}")
        c.drawString(100, 500, f"Is_Automatic: {data.get('is_automatic')}")
        c.drawString(100, 480, f"Is_Closed: {data.get('is_closed')}")
        c.drawString(100, 460, f"Is_Started: {data.get('is_started')}")
        c.drawString(100, 440, f"Started_at: {data.get('started_at')}")
        c.drawString(100, 420, f"Timer_Start: {data.get('timer_start')}")
        c.drawString(100, 400, f"Timer_started_at: {data.get('timer_started_at')}")
        c.drawString(100, 380, f"Timer_Tick_In_Seconds: {data.get('timer_tick_in_seconds')}")
        c.drawString(100, 360, f"Updated_at: {data.get('updated_at')}")
       



        # Add more fields as necessary
        
        c.save()
        
        return bid_pdf_filename
    
    return None



def download_bid_pdf(request, document_id):
    bid_pdf_filename = generate_bid_pdf(document_id)

    if bid_pdf_filename:
        try:
            with open(bid_pdf_filename, 'rb') as file:
                response = HttpResponse(file.read(), content_type='application/pdf')
                response['Content-Disposition'] = 'attachment; filename=' + os.path.basename(bid_pdf_filename)
                print("Response content size:", len(response.content))  # Add this line
                return response
        except FileNotFoundError:
            return HttpResponse("PDF file not found.")
        except Exception as e:
            return HttpResponse("Error generating PDF: " + str(e))

    return HttpResponse("PDF generation failed.") 
# End Bid_Page PDF Download

#Start Bid_Page Excel Sheet Download


def download_bid_excel(request, document_id):
    # Fetch data from Firestore or any other source
    db = firestore.Client()
    doc_ref = db.collection('bids').document(document_id)
    user_doc = doc_ref.get()
    user_data = user_doc.to_dict()

    # Create a new workbook
    workbook = Workbook()

    # Create a new sheet in the workbook
    sheet = workbook.active

    # Add header row
    headers = [
        'Bid_Amount', 
        'Bid Date & Time', 
        'Nth_Bid', 
        'Chit_ID', 
        'Winner_Phone_no',
        'Number_OF_Bells',
        'Bid_at',
        'Ended_at',
        'Is_Active',
        'Is_Agent_bid',
        'Is_Automatic',
        'Is_Closed',
        'Is_Started',
        'Started_at',
        'Timer_Start',
        'Timer_started_at',
        'Timer_Tick_In_Seconds',
        'Updated_at',
        ]  # Add more fields as per your Firestore document structure
    sheet.append(headers)

    # Add data row
#    bid_date = user_data['bid_date']
#    
#    if bid_date.tzinfo is not None:
#        bid_date = bid_date.astimezone(None)
#    bid_date_str = bid_date.strftime('%Y-%m-%d %H:%M:%S')
#    row_data = [
#        user_data['bid_amount'], 
#        bid_date_str, user_data['nth_bid'], 
#        user_data['chit_id'], 
#        user_data['winner_phone_no'],
#    ]

    row_data =[
        user_data['bid_amount'],
        user_data['bid_date'],
        user_data['nth_bid'],
        user_data['chit_id'],
        user_data['winner_phone_no'],
        user_data['num_of_bells'],
        user_data['bid_at'],
        user_data['ended_at'],
        user_data['is_active'],
        user_data['is_agent_bid'],
        user_data['is_automatic'],
        user_data['is_closed'],
        user_data['is_started'],
        user_data['started_at'],
        user_data['timer_start'],
        user_data['timer_started_at'],
        user_data['timer_tick_in_seconds'],
        user_data['updated_at'],
        ]

    row_data = [str(value) if not isinstance(value, list) else ', '.join(value) for value in row_data]

    sheet.append(row_data)

    # Set the response headers for downloading the file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Bid_data.xlsx'


    # Save the workbook to the response
    workbook.save(response)

    return response

# End Bid_Page Excel Sheet Download


# Start Chit Status Active, Completed, Upcoming 
# Start Bid Status function
def bidStatus(request):
    db = firestore.Client()
    # Replace 'your_collection_name' with the actual name of your Firestore collection.
    collection_ref = db.collection('bids')

    if request.method =='GET':
        activeBtn = request.GET.get('activeBtn')
        completedButton = request.GET.get('completedButton')
        upcomingBtn = request.GET.get('upcomingBtn')

        if activeBtn == 'true':
            query = collection_ref.where('is_active', '==', True).stream()
            data_list = [doc.to_dict() for doc in query]

            if not data_list:
                activeErrorMessage = 'Active bids are not available...!'
                return render(request, 'statuscom.html', {'activeErrorMessage': activeErrorMessage})
            
            return render(request, 'statuscom.html', {'data_list': data_list})


        elif completedButton == 'true':
            # Query documents where 'is_closed' field is equal to True.
            query = collection_ref.where('is_closed', '==', True).stream()
            data_list = [doc.to_dict() for doc in query]

            return render(request, 'statuscom.html', {'data_list':data_list})

        elif upcomingBtn == 'false':
            query = collection_ref.where('is_started', '==', False ).where('is_active', '==', False).where('is_closed', '==', False).stream()
            data_list = [doc.to_dict() for doc in query]

            return render(request, 'statuscom.html', {'data_list':data_list})
        else:
           return HttpResponse('Invalid Request...')

    return HttpResponseBadRequest("Invalid request")
# End Bid Status function
# End Chit Status Active, Completed, Upcoming 

# End Bids Page Related codes here...



# Start Chit Page Related Codes Here...

# Read table data Start
def chit_page(request):
    db = firestore.Client()
    docs = db.collection('chits').get()

    chit_data = []

    for doc in docs:
        chit = doc.to_dict()
        chit['id'] = doc.id
        chit_data.append(chit)
    return render(request,'chits.html',{'chit_data':chit_data})

# Start View_Chit Details... 
def chit_details(request,document_id):
    db = firestore.Client()
    get_chitData = db.collection('chits').document(document_id)
    chit = get_chitData.get().to_dict()

    if request.method == 'POST':
        name = request.POST.get('name')
        description = request.POST.get('description')
        agent_commission_type = request.POST.get('agent_commission_type')
        bid_start_amount = request.POST.get('bid_start_amount')
        created_at = request.POST.get('created_at')
        agent_commission_percent = request.POST.get('agent_commission_percent')
        agent_phone_no = request.POST.get('agent_phone_no')
        bid_difference_amount = request.POST.get('bid_difference_amount')
        clients = request.POST.get('clients')
        has_active_bid = request.POST.get('has_active_bid')
        is_bid_started_by_agent = request.POST.get('is_bid_started_by_agent')
        is_closed = request.POST.get('is_closed')
        lapse_in_days = request.POST.get('lapse_in_days')
        last_bid_rebate_value = request.POST.get('last_bid_rebate_value')
        member_commission_type = request.POST.get('member_commission_type')
        next_bid_date = request.POST.get('next_bid_date')
        no_of_clients = request.POST.get('no_of_clients')
        no_of_closed_bids = request.POST.get('no_of_closed_bids')
        payment_grace_period_in_days = request.POST.get('payment_grace_period_in_days')
        payment_grace_period_type = request.POST.get('payment_grace_period_type')
        start_bid_date = request.POST.get('start_bid_date')
        total_amount = request.POST.get('total_amount')
        updated_at = request.POST.get('updated_at')
        winners = request.POST.get('winners')
        past_rebate_values = request.POST.get('past_rebate_values')


        get_chitData.read({
            'name':name,
            'description':description,
            'agent_commission_type':agent_commission_type,
            'bid_start_amount':bid_start_amount,
            'created_at':created_at,
            'agent_commission_percent':agent_commission_percent,
            'agent_phone_no':agent_phone_no,
            'bid_difference_amount':bid_difference_amount,
            'clients':clients,
            'has_active_bid':has_active_bid,
            'is_bid_started_by_agent':is_bid_started_by_agent,
            'is_closed':is_closed,
            'lapse_in_days':lapse_in_days,
            'last_bid_rebate_value':last_bid_rebate_value,
            'member_commission_type':member_commission_type,
            'next_bid_date':next_bid_date,
            'no_of_clients':no_of_clients,
            'no_of_closed_bids':no_of_closed_bids,
            'payment_grace_period_in_days':payment_grace_period_in_days,
            'payment_grace_period_type':payment_grace_period_type,
            'start_bid_date':start_bid_date,
            'total_amount':total_amount,
            'updated_at':updated_at,
            'winners':winners,
            'past_rebate_values':past_rebate_values,
        })
        return redirect('chit_page')
    return render(request, 'chit_details.html',{'document_id':document_id, 'chit':chit})
# End View_Chit Details...

# Chit_Edit_Form Code Start

def chit_edit(request,document_id):
    db = firestore.Client()
    get_chitData = db.collection('chits').document(document_id)
    chit = get_chitData.get().to_dict()

    if request.method == 'POST':
        name = request.POST.get('name')
        description = request.POST.get('description')
        agent_commission_type = request.POST.get('agent_commission_type')
        bid_start_amount = request.POST.get('bid_start_amount')
        created_at = request.POST.get('created_at')
        agent_commission_percent = request.POST.get('agent_commission_percent')
        agent_phone_no = request.POST.get('agent_phone_no')
        bid_difference_amount = request.POST.get('bid_difference_amount')
        clients = request.POST.get('clients')
        has_active_bid = request.POST.get('has_active_bid')
        is_bid_started_by_agent = request.POST.get('is_bid_started_by_agent')
        is_closed = request.POST.get('is_closed')
        lapse_in_days = request.POST.get('lapse_in_days')
        last_bid_rebate_value = request.POST.get('last_bid_rebate_value')
        member_commission_type = request.POST.get('member_commission_type')
        next_bid_date = request.POST.get('next_bid_date')
        no_of_clients = request.POST.get('no_of_clients')
        no_of_closed_bids = request.POST.get('no_of_closed_bids')
        payment_grace_period_in_days = request.POST.get('payment_grace_period_in_days')
        payment_grace_period_type = request.POST.get('payment_grace_period_type')
        start_bid_date = request.POST.get('start_bid_date')
        total_amount = request.POST.get('total_amount')
        updated_at = request.POST.get('updated_at')
        winners = request.POST.get('winners')
        past_rebate_values = request.POST.get('past_rebate_values')


        get_chitData.update({
            'name':name,
            'description':description,
            'agent_commission_type':agent_commission_type,
            'bid_start_amount':bid_start_amount,
            'created_at':created_at,
            'agent_commission_percent':agent_commission_percent,
            'agent_phone_no':agent_phone_no,
            'bid_difference_amount':bid_difference_amount,
            'clients':clients,
            'has_active_bid':has_active_bid,
            'is_bid_started_by_agent':is_bid_started_by_agent,
            'is_closed':is_closed,
            'lapse_in_days':lapse_in_days,
            'last_bid_rebate_value':last_bid_rebate_value,
            'member_commission_type':member_commission_type,
            'next_bid_date':next_bid_date,
            'no_of_clients':no_of_clients,
            'no_of_closed_bids':no_of_closed_bids,
            'payment_grace_period_in_days':payment_grace_period_in_days,
            'payment_grace_period_type':payment_grace_period_type,
            'start_bid_date':start_bid_date,
            'total_amount':total_amount,
            'updated_at':updated_at,
            'winners':winners,
            'past_rebate_values':past_rebate_values,
        })
        return redirect('chit_page')
    return render(request, 'chit_edit_form.html',{'document_id':document_id, 'chit':chit})

# Chit_Edit_Form Code End

# Read table data End...


# Chit Create Page Start...
def chit_create_page(request):
    db = firestore.Client()
    get_chitData = db.collection('chits').document()
    chit = get_chitData.get().to_dict()

    if request.method == 'POST':
        name = request.POST.get('name')
        description = request.POST.get('description')
        agent_commission_type = request.POST.get('agent_commission_type')
        bid_start_amount = request.POST.get('bid_start_amount')
        created_at = request.POST.get('created_at')
        agent_commission_percent = request.POST.get('agent_commission_percent')
        agent_phone_no = request.POST.get('agent_phone_no')
        bid_difference_amount = request.POST.get('bid_difference_amount')
        clients = request.POST.getlist('clients')
        has_active_bid = request.POST.get('has_active_bid')
        is_bid_started_by_agent = request.POST.get('is_bid_started_by_agent')
        is_closed = request.POST.get('is_closed')
        lapse_in_days = request.POST.get('lapse_in_days')
        last_bid_rebate_value = request.POST.get('last_bid_rebate_value')
        member_commission_type = request.POST.get('member_commission_type')
        next_bid_date = request.POST.get('next_bid_date')
        no_of_clients = request.POST.get('no_of_clients')
        no_of_closed_bids = request.POST.get('no_of_closed_bids')
        payment_grace_period_in_days = request.POST.get('payment_grace_period_in_days')
        payment_grace_period_type = request.POST.get('payment_grace_period_type')
        start_bid_date = request.POST.get('start_bid_date')
        total_amount = request.POST.get('total_amount')
        updated_at = request.POST.get('updated_at')
        winners = request.POST.get('winners')
        past_rebate_values = request.POST.get('past_rebate_values')


        get_chitData.set({
            'name':name,
            'description':description,
            'agent_commission_type':agent_commission_type,
            'bid_start_amount':bid_start_amount,
            'created_at':created_at,
            'agent_commission_percent':agent_commission_percent,
            'agent_phone_no':agent_phone_no,
            'bid_difference_amount':bid_difference_amount,
            'clients':clients,
            'has_active_bid':has_active_bid,
            'is_bid_started_by_agent':is_bid_started_by_agent,
            'is_closed':is_closed,
            'lapse_in_days':lapse_in_days,
            'last_bid_rebate_value':last_bid_rebate_value,
            'member_commission_type':member_commission_type,
            'next_bid_date':next_bid_date,
            'no_of_clients':no_of_clients,
            'no_of_closed_bids':no_of_closed_bids,
            'payment_grace_period_in_days':payment_grace_period_in_days,
            'payment_grace_period_type':payment_grace_period_type,
            'start_bid_date':start_bid_date,
            'total_amount':total_amount,
            'updated_at':updated_at,
            'winners':winners,
            'past_rebate_values':past_rebate_values,
        })
        return redirect('chit_page')
    return render(request, 'chit_add_form.html',{'chit':chit})

# Chit Create Page End...

# Chit PDF Download Function Start

def generate_chit_pdf(document_id):
    print( document_id,  "Generating PDF...")
    db = firestore.Client()
    doc_ref = db.collection('chits').document(document_id)
    chit_users = doc_ref.get()
    
    
   
    if chit_users.exists:
        data = chit_users.to_dict()
        chit_pdf_filename = 'Chit_Data.pdf'
        chit_pdf_path = os.path.abspath(chit_pdf_filename)
        print("PDF file path:", chit_pdf_path)
        # Generate PDF
      
        c = canvas.Canvas(chit_pdf_filename, pagesize=letter)
        
        # Write data to PDF
        c.drawString(100, 700, f"Name: {data.get('name')}")
        c.drawString(100, 680, f"Description: {data.get('description')}")
        c.drawString(100, 660, f"Agent_Commission_type: {data.get('agent_commission_type')}")
        c.drawString(100, 640, f"Bid_Start_Amount: {data.get('bid_start_amount')}")
        c.drawString(100, 620, f"Created_At: {data.get('created_at')}")
        c.drawString(100, 600, f"Agent_Phone_No: {data.get('agent_phone_no')}")
        c.drawString(100, 580, f"Bid_Difference_Amount: {data.get('bid_difference_amount')}")
        c.drawString(100, 560, f"Clients: {data.get('clients')}")
        c.drawString(100, 540, f"Has_Active_Bid: {data.get('has_active_bid')}")
        c.drawString(100, 520, f"Is_Bid_Started_By_Agent: {data.get('is_bid_started_by_agent')}")
        c.drawString(100, 500, f"Is_Closed: {data.get('is_closed')}")
        c.drawString(100, 480, f"Lapse_In_Days: {data.get('lapse_in_days')}")
        c.drawString(100, 460, f"Last_Bid_Rebate_Value: {data.get('last_bid_rebate_value')}")
        c.drawString(100, 440, f"Member_Commission_Type: {data.get('member_commission_type')}")
        c.drawString(100, 420, f"Next_Bid_Date: {data.get('next_bid_date')}")
        c.drawString(100, 400, f"No_Of_Closed_Bids: {data.get('no_of_closed_bids')}")
        c.drawString(100, 380, f"Payment_Grace_Period_In_Days: {data.get('payment_grace_period_in_days')}")
        c.drawString(100, 360, f"Payment_Grace_Period_Type: {data.get('payment_grace_period_type')}")
        c.drawString(100, 340, f"Start_Bid_Date: {data.get('start_bid_date')}")
        c.drawString(100, 320, f"Updated_At: {data.get('updated_at')}")
        c.drawString(100, 300, f"Winners: {data.get('winners')}")
        c.drawString(100, 280, f"Past_Rebate_Values: {data.get('past_rebate_values')}")
        c.drawString(100, 260, f"Agent_Commission_Percent: {data.get('agent_commission_percent')}")
        c.drawString(100, 240, f"Total Amount: {data.get('total_amount')}")
        c.drawString(100, 220, f"Number Of Clients: {data.get('no_of_clients')}")


      
       
        # Add more fields as necessary
        
        c.save()
        
        return chit_pdf_filename
    
    return None



def download_chit_pdf(request, document_id):
    chit_pdf_filename = generate_chit_pdf(document_id)

    if chit_pdf_filename:
        try:
            with open(chit_pdf_filename, 'rb') as file:
                response = HttpResponse(file.read(), content_type='application/pdf')
                response['Content-Disposition'] = 'attachment; filename=' + os.path.basename(chit_pdf_filename)
                print("Response content size:", len(response.content))  # Add this line
                return response
        except FileNotFoundError:
            return HttpResponse("PDF file not found.")
        except Exception as e:
            return HttpResponse("Error generating PDF: " + str(e))

    return HttpResponse("PDF generation failed.") 

# Chit PDF Download Function End


# Chit EXCEL Download Function Start


def download_chit_excel(request, document_id):
    # Fetch data from Firestore or any other source
    db = firestore.Client()
    doc_ref = db.collection('chits').document(document_id)
    chit_doc = doc_ref.get()
    chit_data = chit_doc.to_dict()

    # Create a new workbook
    workbook = Workbook()

    # Create a new sheet in the workbook
    sheet = workbook.active

    # Add header row
    headers = ['Name',
               'Description',
               'Agent_Commission_type',
               'Bid_Start_Amount',
               'Created_At',
               'Agent_Phone_No',
               'Bid_Difference_Amount',
               'Clients',
               'Has_Active_Bid',
               'Is_Bid_Started_By_Agent',
               'Is_Closed',
               'Lapse_In_Days',
               'Last_Bid_Rebate_Value',
               'Member_Commission_Type',
               'Next_Bid_Date',
               'No Of Clients',
               'No_Of_Closed_Bids',
               'Payment_Grace_Period_In_Days',
               'Payment_Grace_Period_Type',
               'Start_Bid_Date',
               'Total Amount',
               'Updated_At',
               'Winners',
               'Past_Rebate_Values',
               'Agent_Commission_Percent',
    ]
    sheet.append(headers)

    # Add data row
    row_data = [
        chit_data['name'],
        chit_data['description'],
        chit_data['agent_commission_type'],
        chit_data['bid_start_amount'],
        chit_data['created_at'],
        chit_data['agent_phone_no'],
        chit_data['bid_difference_amount'],
        chit_data['clients'],
        chit_data['has_active_bid'],
        chit_data['is_bid_started_by_agent'],
        chit_data['is_closed'],
        chit_data['lapse_in_days'],
        chit_data['last_bid_rebate_value'],
        chit_data['member_commission_type'],
        chit_data['next_bid_date'],
        chit_data['no_of_clients'],
        chit_data['no_of_closed_bids'],
        chit_data['payment_grace_period_in_days'],
        chit_data['payment_grace_period_type'],
        chit_data['start_bid_date'],
        chit_data['total_amount'],
        chit_data['updated_at'],
        chit_data['winners'],
        chit_data['past_rebate_values'],
        chit_data['agent_commission_percent'],
    ]
    
    row_data = [str(value) if not isinstance(value, list) else ', '.join(value) for value in row_data]

    sheet.append(row_data)

    # Set the response headers for downloading the file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Chit_data.xlsx'

    # Save the workbook to the response
    workbook.save(response)

    return response

# Chit EXCEL Download Function Start


# Chits Delete function Start...
def chit_delete(request, document_id):
    db = firestore.Client()
    doc_ref = db.collection('chits').document(document_id)
    doc_ref.delete()

    return redirect('chit_page')
# Chits Delete function End...

# End Chit Page Related Codes Here...


# Start Practice code...
def prac(request):
    db = firestore.Client()
    docs = db.collection('bids').get()

    bids = []
    for doc in docs:
        bid = doc.to_dict()
        bid['id'] = doc.id
        bids.append(bid)
    return render (request, 'bids.html', {'bids':bids})

# End Practice code...




